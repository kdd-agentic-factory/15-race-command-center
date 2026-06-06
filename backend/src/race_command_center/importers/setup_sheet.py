"""Digital setup-sheet importer — Spec §8 (digitalización de tablas de reglajes).

Parses the team's improved bike-configuration template
(``Mejora de Hoja de Configuración de Moto.xlsx``) — the digital version of the
Aspar paper setup sheet — into a structured, comparable setup record.

Sheet layout::

    CATEGORÍA | PARÁMETRO | SETTING 1 | SETTING 2 | SETTING 3
    GENERAL   | Circuito / Fecha      | …
    FORK      | Muelles (Springs)     | …
    SHOCK     | Compresión (Alta/Baja)| …
    …

Column A introduces a category (carried down for following rows); column B is
the parameter; the remaining columns are independent setting variants that
engineers compare side by side.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_HEADER_TOKENS = {"CATEGORÍA", "CATEGORIA", "PARÁMETRO", "PARAMETRO"}


def _norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_setup_sheet(path: str | Path) -> dict[str, Any]:
    """Parse the digital setup-sheet workbook into structured categories.

    Returns
    -------
    dict with ``setting_labels`` (the SETTING column headers) and ``categories``,
    each ``{"name", "parameters": [{"parameter", "settings": {label: value}}]}``.

    Raises
    ------
    FileNotFoundError, ValueError
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Setup-sheet workbook not found: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [[ _norm(c) for c in row ] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    # Locate the header row (contains CATEGORÍA / PARÁMETRO).
    header_idx = None
    for i, row in enumerate(rows):
        upper = {c.upper() for c in row if c}
        if upper & _HEADER_TOKENS:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"No header row (CATEGORÍA/PARÁMETRO) in {path.name}")

    header = rows[header_idx]
    # Setting columns are everything after the PARÁMETRO column (index 1).
    setting_labels = [h for h in header[2:] if h] or ["SETTING 1"]

    categories: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    current_name = ""
    for row in rows[header_idx + 1 :]:
        if not any(row):
            continue
        cat, param = row[0], row[1]
        values = row[2 : 2 + len(setting_labels)]
        if cat:
            current_name = cat
        if not param:
            continue
        if current is None or current["name"] != current_name:
            current = {"name": current_name or "GENERAL", "parameters": []}
            categories.append(current)
        current["parameters"].append({
            "parameter": param,
            "settings": {label: (values[j] if j < len(values) else "") for j, label in enumerate(setting_labels)},
        })

    if not categories:
        raise ValueError(f"No parameters parsed from {path.name}")

    param_count = sum(len(c["parameters"]) for c in categories)
    logger.info("Parsed setup sheet %s: %d categories, %d parameters",
                path.name, len(categories), param_count)
    return {
        "format": "setup-sheet",
        "source_file": path.name,
        "setting_labels": setting_labels,
        "category_count": len(categories),
        "parameter_count": param_count,
        "categories": categories,
    }
