"""Engine-Brake (AEB) map importer — Spec §8.2 ECU / Freno Motor.

Parses the 2D-style **EB Tool** Excel workbook (``EB_Tool_<model>.xlsm``) that
defines the engine-brake / AEB retention maps used by the ECU. The workbook has
one sheet per gear (``GEAR 1`` … ``GEAR 6``) plus a consolidated ``DataBase``
sheet. Each gear sheet is a 2D lookup table:

    row index   = engine RPM breakpoints     (column ``BP``: 0, 1000, … 9000+)
    col index   = brake-position breakpoints  (header row: 0, 7, 13, … 59)
    cell value  = engine-brake (AEB) retention amount

This module turns that into structured, JSON-serialisable maps so the platform
can ingest ECU parameter files and reason about engine-brake by gear/RPM/brake.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_MAX_HEADER_SCAN = 20


def _to_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_gear_sheet(ws) -> dict[str, Any] | None:
    """Parse a single ``GEAR n`` sheet into an engine-brake map."""
    # The header row is the one whose column B holds the literal "BP".
    header_row = None
    for r in range(1, _MAX_HEADER_SCAN + 1):
        if str(ws.cell(r, 2).value).strip().upper() == "BP":
            header_row = r
            break
    if header_row is None:
        return None

    eb_type = ws.cell(header_row, 1).value  # e.g. "EBb" / "EBi"

    # Brake-position breakpoints: column C onward on the header row.
    brake_positions: list[float] = []
    cols: list[int] = []
    col = 3
    while col <= ws.max_column:
        v = _to_float(ws.cell(header_row, col).value)
        if v is None:
            break
        brake_positions.append(v)
        cols.append(col)
        col += 1
    if not cols:
        return None

    # Data rows: RPM breakpoint in column B, AEB values across the brake columns.
    rpm_breakpoints: list[float] = []
    grid: list[list[float]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        rpm = _to_float(ws.cell(r, 2).value)
        if rpm is None:
            break
        rpm_breakpoints.append(rpm)
        grid.append([_to_float(ws.cell(r, c).value) or 0.0 for c in cols])

    return {
        "eb_type": str(eb_type).strip() if eb_type is not None else None,
        "brake_positions": brake_positions,
        "rpm_breakpoints": rpm_breakpoints,
        "map": grid,
        "nonzero_cells": sum(1 for row in grid for v in row if v != 0.0),
    }


def parse_engine_brake_workbook(path: str | Path) -> dict[str, Any]:
    """Parse an EB Tool ``.xlsm`` into structured engine-brake maps per gear.

    Raises
    ------
    FileNotFoundError
        If the file does not exist.
    ValueError
        If no gear sheets could be parsed.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"EB Tool workbook not found: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    gears: list[dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.strip().upper().startswith("GEAR"):
                continue
            gear_no = int("".join(ch for ch in sheet_name if ch.isdigit()) or 0)
            parsed = _parse_gear_sheet(wb[sheet_name])
            if parsed is None:
                continue
            parsed["gear"] = gear_no
            gears.append(parsed)
    finally:
        # read_only mode keeps the file handle open until close(); on Windows a
        # caller cannot unlink the temp upload otherwise.
        wb.close()

    if not gears:
        raise ValueError(f"No gear sheets parsed from {path.name}")

    gears.sort(key=lambda g: g["gear"])
    model = path.stem.replace("EB_Tool_", "").replace("EB_Tool", "") or path.stem
    logger.info("Parsed engine-brake workbook %s: %d gears", path.name, len(gears))
    return {
        "format": "2d-eb-tool",
        "source_file": path.name,
        "model": model,
        "gear_count": len(gears),
        "gears": gears,
    }
