"""Tests for the Engine-Brake (AEB) map importer — Spec §8.2."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from race_command_center.importers.engine_brake import parse_engine_brake_workbook


def _build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for gear in (1, 2):
        ws = wb.create_sheet(f"GEAR {gear}")
        ws["A1"] = "EB Curve Select"
        # header row: A2 = type, B2 = BP, then brake positions
        ws["A2"] = "EBb" if gear == 1 else "EBi"
        ws["B2"] = "BP"
        for col, pos in zip("CDE", (0, 7, 13)):
            ws[f"{col}2"] = pos
        # data rows: B = RPM, C..E = AEB values
        rows = [(0, 0, 0, 0), (8750, 2.15, 0.45, 0.0), (9000, 4.2, 1.25, 0.2)]
        for i, (rpm, c, d, e) in enumerate(rows, start=3):
            ws[f"B{i}"] = rpm
            ws[f"C{i}"], ws[f"D{i}"], ws[f"E{i}"] = c, d, e
    # an unrelated sheet must be ignored
    wb.create_sheet("DataBase")
    return wb


@pytest.fixture()
def eb_xlsm(tmp_path: Path) -> Path:
    p = tmp_path / "EB_Tool_TEST.xlsm"
    _build_workbook().save(p)
    return p


def test_parse_gears_and_values(eb_xlsm: Path):
    d = parse_engine_brake_workbook(eb_xlsm)
    assert d["model"] == "TEST"
    assert d["gear_count"] == 2
    g1, g2 = d["gears"]
    assert g1["gear"] == 1 and g1["eb_type"] == "EBb"
    assert g2["gear"] == 2 and g2["eb_type"] == "EBi"
    assert g1["brake_positions"] == [0.0, 7.0, 13.0]
    assert g1["rpm_breakpoints"] == [0.0, 8750.0, 9000.0]
    # value grid (last row = 9000 rpm)
    assert g1["map"][2] == [4.2, 1.25, 0.2]
    assert g1["nonzero_cells"] == 5


def test_rejects_workbook_without_gears(tmp_path: Path):
    wb = Workbook()
    wb.active.title = "Summary"
    p = tmp_path / "EB_Tool_EMPTY.xlsm"
    wb.save(p)
    with pytest.raises(ValueError):
        parse_engine_brake_workbook(p)


def test_missing_file():
    with pytest.raises(FileNotFoundError):
        parse_engine_brake_workbook("/nope/EB_Tool.xlsm")


def test_import_endpoint(eb_xlsm: Path):
    from fastapi.testclient import TestClient

    from race_command_center.main import app

    client = TestClient(app)
    with open(eb_xlsm, "rb") as fh:
        buf = io.BytesIO(fh.read())
    r = client.post(
        "/api/v1/engine-brake/import",
        files={"file": ("EB_Tool_TEST.xlsm", buf, "application/vnd.ms-excel.sheet.macroEnabled.12")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["gear_count"] == 2
    assert body["gears"][0]["eb_type"] == "EBb"
